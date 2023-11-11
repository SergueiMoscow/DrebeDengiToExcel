import openpyxl
import openpyxl as xl
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import numbers


def format_xlsx(input_file: str, sheet_name: str, operation: str, value: any, **kwargs):
    workbook = xl.load_workbook(input_file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    if operation == 'width':
        sheet.column_dimensions[kwargs.get('column')].width = value
    workbook.save()


def reformat_xlsx(file_name: str, dest_file_name: str):
    workbook = xl.load_workbook(file_name)
    sheet = workbook['expense']
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['G'].width = 15
    workbook.save(filename=dest_file_name)


def transform_excel_column_to_num(excel_file_name, sheet_name, column_name) -> None:
    """
    Меняет запятые на точки, убирает одинарные кавычки и устанавливает формат числа с 2-мя десятичными
    для ячейки column_name листа sheet_name файла excel_file_name
    :param excel_file_name:
    :param sheet_name:
    :param column_name:
    :return:
    """
    workbook = openpyxl.load_workbook(excel_file_name)

    # Получить лист по имени
    sheet = workbook[sheet_name]

    # Найти индекс колонки
    column_index = None
    for i, cell in enumerate(sheet[1]):
        if cell.value == column_name:
            column_index = i + 1
            break
    if column_index is None:
        print(f"Column '{column_name}' not found.")
        return

    for row in sheet.iter_rows(min_col=column_index, max_col=column_index, min_row=2):
        cell = row[0]
        if cell.value is not None:
            # Замена запятой на точку и удаление одинарной кавычки
            value = str(cell.value).replace(",", ".").lstrip("'")

            try:
                numeric_value = round(float(value), 2)
            except ValueError:
                print(f"Conversion error: {cell.value}")
                continue

            cell.value = numeric_value
            cell.number_format = numbers.FORMAT_NUMBER_00

    workbook.save(excel_file_name)
    print(f"File '{excel_file_name}' has been updated.")
