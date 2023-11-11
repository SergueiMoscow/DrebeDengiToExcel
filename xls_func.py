import os

import openpyxl
import pandas as pd
import xlsxwriter
import xlwings
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.filters import FilterColumn

MAP_EXPENSES = {
    'expense': ['Категория', 'Член семьи', 'Дата', 'Комментарий', 'Валюта', '№ Категории', 'Сумма', 'Счёт'],
    'income': ['Категория', 'Член семьи', 'Дата', 'Комментарий', 'Валюта', '№ Категории', 'Сумма', 'Счёт'],
    'transfer': ['Откуда', 'Куда', 'Сумма', 'Член семьи', 'Дата', 'Комментарий', 'Валюта', '№ Откуда', '№ Куда'],
    'currency': ['Название', 'курс', 'id'],
    'expense_category': ['Название', 'id', 'parent'],
    'income_category': ['Название', 'id', 'parent'],
    'account_category': ['Название', 'id', 'parent'],
}


def convert_csv_to_xlsx(csv_file_path) -> str:
    """
    Принимает csv, конвертирует в xls, и возвращает имя xls файла
    :param csv_file_path:
    :return: xls_file_path
    """
    # Проверяем, существует ли csv файл
    if not os.path.exists(csv_file_path):
        print(f'Файл {csv_file_path} не найден')
        return ''

    # Читаем данные из csv файла
    df = pd.read_csv(csv_file_path)

    # Получаем путь без расширения
    file_path_without_extension = os.path.splitext(csv_file_path)[0]

    # Создаем новый путь с расширением .xlsx
    xlsx_file_path = file_path_without_extension + '.xlsx'

    # Записываем данные в xlsx файл
    df.to_excel(xlsx_file_path, index=False, engine='openpyxl')

    print(f'Файл {xlsx_file_path} успешно создан')
    return xlsx_file_path


def csv_to_ods(csv_filepath) -> str:
    # имя файла без расширения
    file_base_name = os.path.splitext(csv_filepath)[0]
    # путь к новому файлу ods
    ods_filepath = f"{file_base_name}.ods"

    # чтение CSV-файла
    df = pd.read_csv(csv_filepath)

    # сохранение файла как ODS
    df.to_excel(ods_filepath, engine="openpyxl", index=False)
    return ods_filepath


def import_csv_to_xlsx(csv_filename, xlsx_filename, sheetname):
    # Считывание CSV файла
    # if sheetname in MAP_EXPENSES:
    #     df = pd.read_csv(csv_filename, delimiter=';', names=MAP_EXPENSES[sheetname])
    # else:

    df = pd.read_csv(csv_filename, delimiter=';', header=None)

    if sheetname in MAP_EXPENSES:
        df.columns = MAP_EXPENSES[sheetname]


    if os.path.exists(xlsx_filename):
        # Если xlsx файл уже существует
        book = load_workbook(xlsx_filename)

        # Проверяем, есть ли лист в xlsx файле
        if sheetname in book.sheetnames:
            print('Error: sheetname {} already exists in {}'.format(sheetname, xlsx_filename))
        else:
            with pd.ExcelWriter(xlsx_filename, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, index=False, sheet_name=sheetname, freeze_panes=(1, 1))
    else:
        # Если xlsx файла не существует, создаем новый
        df.to_excel(xlsx_filename, index=False, engine='openpyxl', sheet_name=sheetname)
        print(f'{xlsx_filename} сохранён')


def get_list_name_from_file_name(file_name):
    # отсекаем начальные числовые символы
    list_name = file_name
    pos = list_name.find('_')
    if pos != -1:
        list_name = list_name[pos + 1:]

    # отсекаем конечные символы после последней точки
    pos = list_name.rfind('.')
    if pos != -1:
        list_name = list_name[:pos]

    return list_name


def remove_sheet_from_xlsx(file_name: str, sheet_name: str):
    book = load_workbook(file_name)
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)
        book.save(file_name)


def set_auto_filter(file_name, sheet_name):
    workbook = load_workbook(file_name)
    writer = pd.ExcelWriter('filter_output.xlsx', engine='openpyxl')
    ws = writer.sheets[sheet_name]  # workbook.get_sheet_by_name(sheet_name)
    ws.auto_filter.ref = 'A:B'
    ws.auto_filter.add_filter_column(1, ['INVALID'], blank=False)
    writer.close()
    workbook.close()


def replace_values_from_another_sheet(
        file_name: str,
        target_sheet: str,
        target_col: str,
        source_sheet: str,
        source_key_col: str,
        source_value_col: str
):
    # Загружаем листы в DataFrames
    df_target = pd.read_excel(file_name, sheet_name=target_sheet)
    df_source = pd.read_excel(file_name, sheet_name=source_sheet)

    # Создаем словарь для замены
    replace_dict = pd.Series(df_source[source_value_col].values, index=df_source[source_key_col]).to_dict()

    # Производим замену в целевом листе
    df_target[target_col] = df_target[target_col].replace(replace_dict)
    remove_sheet_from_xlsx(file_name, target_sheet)

    # Записываем обновленные данные обратно в Excel
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        df_target.to_excel(writer, sheet_name=target_sheet, index=False, freeze_panes=(1, 0))
        ws = writer.sheets[target_sheet]
        ws.auto_filter.ref = 'A:H'

